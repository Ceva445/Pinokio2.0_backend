"""Тести звіту «Historia rejestracji» (XLSX).

Перевіряємо три речі: фільтри звужують вибірку так, як обіцяє форма; файл
містить рівно ті самі рядки, що й підгляд; ім'я файлу — назва звіту плюс дата
генерації.
"""
from datetime import date, datetime, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook

from models.db_device import DeviceDB, DeviceType
from models.db_employee import EmployeeDB
from models.db_transaction import TransactionDB, TransactionType
from models.db_user import UserDB
from routers.admin.reports import (
    COLUMNS,
    REPORT_TZ,
    build_workbook,
    registration_report_options,
    registration_report_preview,
    registration_report_xlsx,
    report_file_name,
)

pytestmark = pytest.mark.asyncio


def _moment(day: int, hour: int = 12) -> datetime:
    """Час у тій самій зоні, у якій звіт рахує межі дат."""
    return datetime(2026, 3, day, hour, 0, tzinfo=REPORT_TZ)


@pytest.fixture
async def history(db_session):
    """Дві людини з різних відділів, два пристрої, чотири реєстрації в різні дні."""
    manager = UserDB(
        first_name="Ola", last_name="Zima", username="ozima",
        password_hash="x", role="admin", is_active=True,
    )
    anna = EmployeeDB(
        last_name="Nowak", first_name="Anna", rfid="rfid-anna",
        company="ACME", wms_login="A-NOWAK", department="STOCK",
    )
    bartek = EmployeeDB(
        last_name="Wojcik", first_name="Bartek", rfid="rfid-bartek",
        company="ACME", wms_login="B-WOJCIK", department="ECOM",
    )
    db_session.add_all([manager, anna, bartek])
    await db_session.commit()

    scanner = DeviceDB(
        name="TERM003", rfid="rfid-t3", serial_number="sn-t3", type=DeviceType.scanner
    )
    printer = DeviceDB(
        name="ZEBRAMOB44506", rfid="rfid-z6", serial_number="sn-z6", type=DeviceType.printer
    )
    db_session.add_all([scanner, printer])
    await db_session.commit()

    db_session.add_all([
        TransactionDB(timestamp=_moment(10), type=TransactionType.registered,
                      employee_id=anna.id, device_id=scanner.id, manager_id=manager.id),
        TransactionDB(timestamp=_moment(11), type=TransactionType.unregistered,
                      employee_id=anna.id, device_id=scanner.id),
        TransactionDB(timestamp=_moment(12), type=TransactionType.registered,
                      employee_id=bartek.id, device_id=printer.id),
        TransactionDB(timestamp=_moment(13), type=TransactionType.registered,
                      employee_id=None, device_id=printer.id),
    ])
    await db_session.commit()

    return {"anna": anna, "bartek": bartek, "scanner": scanner, "printer": printer}


async def _preview(db, **filters):
    params = {"date_from": None, "date_to": None, "department": None, "device_ids": []}
    params.update(filters)
    return await registration_report_preview(db=db, user=None, **params)


# ---------------------------------------------------------------------------
# Фільтри
# ---------------------------------------------------------------------------
async def test_without_filters_report_holds_everything(db_session, history):
    report = await _preview(db_session)

    assert report["total"] == 4
    assert report["columns"] == list(COLUMNS)
    # Найновіше зверху — так само, як на екрані "Rejestracje".
    assert [row[0] for row in report["rows"]] == sorted(
        (row[0] for row in report["rows"]), reverse=True
    )


async def test_date_bounds_are_inclusive(db_session, history):
    """Обидві межі включні: день "do" береться цілком, а не до півночі."""
    report = await _preview(db_session, date_from=date(2026, 3, 11), date_to=date(2026, 3, 12))

    assert report["total"] == 2
    assert all(row[0].startswith(("2026-03-11", "2026-03-12")) for row in report["rows"])


async def test_department_filter_follows_the_employee(db_session, history):
    report = await _preview(db_session, department="STOCK")

    assert report["total"] == 2
    assert all("A-NOWAK" in row[2] for row in report["rows"])


async def test_several_devices_can_be_picked_at_once(db_session, history):
    both = await _preview(
        db_session, device_ids=[history["scanner"].id, history["printer"].id]
    )
    only_scanner = await _preview(db_session, device_ids=[history["scanner"].id])

    assert both["total"] == 4
    assert only_scanner["total"] == 2
    assert {row[3] for row in only_scanner["rows"]} == {"TERM003"}


async def test_filters_stack(db_session, history):
    report = await _preview(
        db_session,
        date_from=date(2026, 3, 10),
        date_to=date(2026, 3, 10),
        department="STOCK",
        device_ids=[history["scanner"].id],
    )

    assert report["total"] == 1
    assert report["rows"][0][1] == "registered"


async def test_missing_employee_and_manager_show_a_dash(db_session, history):
    report = await _preview(db_session, date_from=date(2026, 3, 13))

    assert report["total"] == 1
    row = report["rows"][0]
    assert row[2] == "—"          # rejestracja bez pracownika
    assert row[4] == "—"          # i bez managera
    assert row[3] == "ZEBRAMOB44506"


async def test_manager_is_named_like_on_the_screen(db_session, history):
    report = await _preview(db_session, date_from=date(2026, 3, 10), date_to=date(2026, 3, 10))

    assert report["rows"][0][4] == "Ola Zima (ozima)"


# ---------------------------------------------------------------------------
# Опції фільтрів
# ---------------------------------------------------------------------------
async def test_options_feed_the_dropdowns(db_session, history):
    options = await registration_report_options(db=db_session, user=None)

    assert options["departments"] == ["ECOM", "STOCK"]
    assert [d["name"] for d in options["devices"]] == ["TERM003", "ZEBRAMOB44506"]


# ---------------------------------------------------------------------------
# Сам файл
# ---------------------------------------------------------------------------
async def test_file_name_is_report_name_plus_generation_date():
    assert report_file_name(date(2026, 9, 4)) == "Historia_rejestracji_2026-09-04.xlsx"


async def test_workbook_matches_what_the_preview_showed(db_session, history):
    preview = await _preview(db_session, department="STOCK")
    response = await registration_report_xlsx(
        db=db_session, user=None,
        date_from=None, date_to=None, department="STOCK", device_ids=[],
    )

    sheet = load_workbook(BytesIO(response.body)).active
    rows = list(sheet.values)

    assert rows[0] == COLUMNS
    assert len(rows) - 1 == preview["total"] == 2

    # Дата в клітинці — справжній datetime, щоб Excel міг сортувати й фільтрувати.
    assert isinstance(rows[1][0], datetime)
    assert [r[3] for r in rows[1:]] == [row[3] for row in preview["rows"]]


async def test_workbook_has_headers_even_when_nothing_matches(db_session, history):
    response = await registration_report_xlsx(
        db=db_session, user=None,
        date_from=date(2030, 1, 1), date_to=None, department=None, device_ids=[],
    )

    sheet = load_workbook(BytesIO(response.body)).active
    assert list(sheet.values) == [COLUMNS]


async def test_response_carries_the_file_name(db_session, history):
    response = await registration_report_xlsx(
        db=db_session, user=None,
        date_from=None, date_to=None, department=None, device_ids=[],
    )

    today = datetime.now(REPORT_TZ).date()
    assert report_file_name(today) in response.headers["content-disposition"]
    assert response.media_type.endswith("spreadsheetml.sheet")


async def test_timestamps_are_written_without_a_timezone(db_session, history):
    """openpyxl не вміє писати tz-aware дати — вони мають приїхати вже локальними."""
    sheet = load_workbook(BytesIO(build_workbook(
        [(datetime(2026, 3, 10, 12, 0), "registered", "A-NOWAK", "TERM003", "—")]
    ))).active

    written = list(sheet.values)[1][0]
    assert written.tzinfo is None
    assert written == datetime(2026, 3, 10, 12, 0)


async def test_report_covers_a_span_of_days(db_session, history):
    """Межі беруться від початку першого дня до кінця останнього."""
    span = await _preview(db_session, date_from=date(2026, 3, 10), date_to=date(2026, 3, 13))
    assert span["total"] == 4

    narrower = await _preview(db_session, date_from=_moment(10).date() + timedelta(days=1))
    assert narrower["total"] == 3
