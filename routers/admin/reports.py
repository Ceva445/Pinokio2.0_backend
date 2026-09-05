"""Raporty administracyjne.

Na razie jeden: historia rejestracji do XLSX. Kolumny są takie same jak na
ekranie "Rejestracje" — raport ma być tym samym zestawieniem, tylko bez
stronicowania i z własnymi filtrami (zakres dat, dział, wybrane urządzenia).

Podgląd i plik czytają z jednego zapytania, więc to, co widać na ekranie, jest
tym, co wyląduje w pliku.
"""
from datetime import date, datetime, time
from io import BytesIO
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.dependencies.admin import require_admin
from db.session import get_db
from models.db_device import DeviceDB
from models.db_employee import EmployeeDB
from models.db_transaction import TransactionDB

router = APIRouter(
    prefix="/admin/api/reports",
    tags=["Admin Reports"]
)

REPORT_NAME = "Historia rejestracji"

# Magazyn stoi w Polsce, a app/main.py planuje już maile w tej samej strefie.
# Godziny w pliku mają się zgadzać z tym, co admin widzi w przeglądarce.
REPORT_TZ = ZoneInfo("Europe/Warsaw")

# Kolejność musi odpowiadać tabeli na /admin/transactions.
COLUMNS = ("Data / czas", "Typ", "Pracownik", "Urządzenie", "Manager")
COLUMN_WIDTHS = (22, 16, 34, 24, 30)

PREVIEW_LIMIT = 50


def _filtered(
    date_from: date | None,
    date_to: date | None,
    department: str | None,
    device_ids: list[int] | None,
):
    """Wspólny rdzeń raportu: te same warunki dla podglądu, liczby i pliku."""
    stmt = (
        select(TransactionDB)
        .outerjoin(TransactionDB.employee)
        .join(TransactionDB.device)
    )

    if date_from:
        stmt = stmt.where(
            TransactionDB.timestamp >= datetime.combine(date_from, time.min, tzinfo=REPORT_TZ)
        )

    if date_to:
        # Górna granica obejmuje cały dzień, tak jak filtr na ekranie rejestracji.
        stmt = stmt.where(
            TransactionDB.timestamp <= datetime.combine(date_to, time.max, tzinfo=REPORT_TZ)
        )

    if department:
        stmt = stmt.where(EmployeeDB.department == department)

    if device_ids:
        stmt = stmt.where(DeviceDB.id.in_(device_ids))

    return stmt


def _employee_label(employee) -> str:
    if not employee:
        return "—"
    parts = (employee.wms_login, employee.first_name, employee.last_name)
    return " ".join(p for p in parts if p).strip() or "—"


def _manager_label(manager) -> str:
    if not manager:
        return "—"
    name = " ".join(p for p in (manager.first_name, manager.last_name) if p).strip()
    if not name:
        return manager.username or "—"
    return f"{name} ({manager.username})" if manager.username else name


def _local(moment: datetime | None) -> datetime | None:
    """Znacznik czasu w strefie raportu, bez tzinfo — Excel stref nie zapisuje."""
    if moment is None:
        return None
    if moment.tzinfo is None:
        return moment
    return moment.astimezone(REPORT_TZ).replace(tzinfo=None)


def _row(transaction: TransactionDB) -> tuple:
    return (
        _local(transaction.timestamp),
        transaction.type.value,
        _employee_label(transaction.employee),
        transaction.device.name if transaction.device else "—",
        _manager_label(transaction.manager),
    )


async def _rows(db: AsyncSession, limit: int | None = None, **filters) -> list[tuple]:
    stmt = (
        _filtered(**filters)
        .options(
            joinedload(TransactionDB.employee),
            joinedload(TransactionDB.device),
            joinedload(TransactionDB.manager),
        )
        .order_by(TransactionDB.timestamp.desc())
    )
    if limit is not None:
        stmt = stmt.limit(limit)

    return [_row(t) for t in (await db.execute(stmt)).scalars().all()]


async def _count(db: AsyncSession, **filters) -> int:
    stmt = select(func.count()).select_from(_filtered(**filters).subquery())
    return (await db.execute(stmt)).scalar_one()


def report_file_name(generated_on: date) -> str:
    """Nazwa raportu plus data wygenerowania, np. Historia_rejestracji_2026-09-04."""
    return f"{REPORT_NAME.replace(' ', '_')}_{generated_on.isoformat()}.xlsx"


def build_workbook(rows: list[tuple]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rejestracje"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")

    sheet.append(list(COLUMNS))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        sheet.append(list(row))

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for cell in sheet["A"][1:]:
        cell.number_format = "YYYY-MM-DD HH:MM:SS"

    # Nagłówek zostaje na widoku przy przewijaniu, filtry ułatwiają przeglądanie.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{sheet.max_row}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@router.get("/registrations/options")
async def registration_report_options(
    db: AsyncSession = Depends(get_db),
    user=Depends(require_admin)
):
    """Zawartość filtrów: działy pracowników i lista urządzeń do wyboru."""
    departments = (await db.execute(
        select(EmployeeDB.department)
        .where(EmployeeDB.department.is_not(None))
        .distinct()
        .order_by(EmployeeDB.department)
    )).scalars().all()

    devices = (await db.execute(
        select(DeviceDB).order_by(DeviceDB.name)
    )).scalars().all()

    return {
        "report_name": REPORT_NAME,
        "departments": list(departments),
        "devices": [
            {"id": d.id, "name": d.name, "type": d.type.value, "enabled": d.enabled}
            for d in devices
        ],
    }


@router.get("/registrations")
async def registration_report_preview(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department: str | None = Query(default=None),
    device_ids: list[int] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_admin)
):
    """Pierwsze wiersze raportu plus ich łączna liczba — żeby admin wiedział,
    co pobiera, zanim pobierze."""
    filters = {
        "date_from": date_from,
        "date_to": date_to,
        "department": department,
        "device_ids": device_ids,
    }

    rows = await _rows(db, limit=PREVIEW_LIMIT, **filters)

    return {
        "columns": list(COLUMNS),
        "rows": [
            [value.isoformat(sep=" ", timespec="seconds") if isinstance(value, datetime) else value
             for value in row]
            for row in rows
        ],
        "total": await _count(db, **filters),
        "preview_limit": PREVIEW_LIMIT,
        "file_name": report_file_name(datetime.now(REPORT_TZ).date()),
    }


@router.get("/registrations.xlsx")
async def registration_report_xlsx(
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    department: str | None = Query(default=None),
    device_ids: list[int] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
    user=Depends(require_admin)
):
    """Cała historia rejestracji spełniająca filtry, jako plik XLSX."""
    rows = await _rows(
        db,
        date_from=date_from,
        date_to=date_to,
        department=department,
        device_ids=device_ids,
    )

    file_name = report_file_name(datetime.now(REPORT_TZ).date())

    return Response(
        content=build_workbook(rows),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )
